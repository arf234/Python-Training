import os
import openpyxl as xl
from openpyxl.chart import BarChart, Reference
     #Package.#module      #Classes
      
print("Current Working Directory:", os.getcwd())
os.chdir(r'c:\...')
print("Current Working Directory:", os.getcwd())


def process_workbook(filename):
    wb= xl.load_workbook('transactions.xlsx')
    sheet = wb['Sheet1']
    #cell = sheet['a1']
    #cell = sheet.cell(1, 1)
    #print(cell.value)

    #print(sheet.max_row) #Checking the total rows in the table

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
            min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)